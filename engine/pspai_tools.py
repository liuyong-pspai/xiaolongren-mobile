"""
🐉 小龙人 PSPAI 自定义工具集 v2.0
=================================
全部工具采用统一 def handler(args) 签名。
不管是通过GENE://协议还是/api/tool/exec调用，参数都是字典，工具负责解包。
每个工具独立可运行，零外部依赖（标准库），第三方包走惰性import。
"""
import json
import logging
import os
import re
import sqlite3
import subprocess
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, Any

logger = logging.getLogger(__name__) if __name__ != "__main__" else __import__('logging').getLogger(__name__)


# ══════════════════════════════════════════════
# 所有工具采用统一签名: def tool(args: dict) -> str
# "args" 是从GENE://或/api/tool/exec传入的参数字典
# ══════════════════════════════════════════════

# ---- 工具1: git_operations ----
def git_operations(args: Dict[str, Any]) -> str:
    """Git操作: status/add/commit/push/pull/log/diff"""
    action = args.get("action", "")
    repo_path = args.get("repo_path", ".")
    message = args.get("message", "")
    branch = args.get("branch", "")
    remote = args.get("remote", "origin")
    try:
        cmds = {
            "status": ["git", "status", "--short"],
            "add": ["git", "add", "."],
            "branch": ["git", "branch", "-a"],
            "log": ["git", "log", "--oneline", "-10"],
            "diff": ["git", "diff", "--stat"],
        }
        if action in cmds:
            r = subprocess.run(cmds[action], capture_output=True, text=True, cwd=repo_path, timeout=30)
            return r.stdout or r.stderr or "完成"
        elif action == "commit":
            if not message:
                return "❌ 需提供message"
            r = subprocess.run(["git", "commit", "-m", message], capture_output=True, text=True, cwd=repo_path, timeout=30)
            return r.stdout or r.stderr
        elif action == "push":
            r = subprocess.run(["git", "push", remote, branch or "HEAD"], capture_output=True, text=True, cwd=repo_path, timeout=120)
            return r.stdout or r.stderr
        elif action == "pull":
            args_list = ["git", "pull"]
            if branch:
                args_list += [remote, branch]
            r = subprocess.run(args_list, capture_output=True, text=True, cwd=repo_path, timeout=120)
            return r.stdout or r.stderr
        elif action == "checkout":
            if not branch:
                return "❌ 需提供branch"
            r = subprocess.run(["git", "checkout", branch], capture_output=True, text=True, cwd=repo_path, timeout=30)
            return r.stdout or r.stderr
        return f"❌ 不支持的Git操作: {action}"
    except subprocess.TimeoutExpired:
        return "❌ 超时"
    except FileNotFoundError:
        return "❌ 不是Git仓库"
    except Exception as e:
        return f"❌ Git失败: {e}"


# ---- 工具2: sql_query ----
def sql_query(args: Dict[str, Any]) -> str:
    """执行SQL查询（SQLite）"""
    query = args.get("query", "")
    db_path = args.get("db_path", str(Path.home() / ".pspai" / "state.db"))
    if not query:
        return "❌ 需提供SQL语句"
    try:
        db_file = Path(db_path)
        if not db_file.exists():
            return f"❌ 数据库不存在: {db_path}"
        conn = sqlite3.connect(str(db_file))
        conn.row_factory = sqlite3.Row
        cur = conn.execute(query)
        if query.strip().upper().startswith(("SELECT", "PRAGMA")):
            rows = cur.fetchall()
            if not rows:
                return "空结果"
            cols = [d[0] for d in cur.description]
            result = " | ".join(cols) + "\n" + "-" * 60
            for i, row in enumerate(rows):
                if i >= 50:
                    break
                vals = [str(row[c])[:30] if row[c] is not None else "" for c in cols]
                result += "\n" + " | ".join(vals)
            if len(rows) > 50:
                result += f"\n...共{len(rows)}行"
            return result
        else:
            conn.commit()
            return f"✅ 影响{cur.rowcount}行"
    except sqlite3.Error as e:
        return f"❌ SQL错误: {e}"
    except Exception as e:
        return f"❌ 失败: {e}"


# ---- 工具3: csv_read ----
def csv_read(args: Dict[str, Any]) -> str:
    """读取CSV文件"""
    path = args.get("path", "")
    max_rows = int(args.get("max_rows", 20))
    delimiter = args.get("delimiter", ",")
    if not path:
        return "❌ 需提供文件路径"
    try:
        import csv
        p = Path(path)
        if not p.exists():
            return f"❌ 文件不存在: {path}"
        rows = []
        with open(p, newline="", encoding="utf-8", errors="replace") as f:
            for i, row in enumerate(csv.reader(f, delimiter=delimiter)):
                if i > max_rows:
                    break
                rows.append(row)
        if not rows:
            return "空文件"
        r = " | ".join(c[:25] for c in rows[0]) + "\n" + "-" * 60
        for row in rows[1:]:
            r += "\n" + " | ".join((c[:25] if len(c) > 25 else c) for c in row)
        return r + f"\n--- 共{len(rows)-1}行 ---"
    except ImportError:
        return "❌ 需要csv模块（标准库应有）"
    except Exception as e:
        return f"❌ CSV错误: {e}"


# ---- 工具4: excel_read ----
def excel_read(args: Dict[str, Any]) -> str:
    """读取Excel文件"""
    path = args.get("path", "")
    sheet = args.get("sheet", "")
    max_rows = int(args.get("max_rows", 20))
    if not path:
        return "❌ 需提供文件路径"
    try:
        import openpyxl
        p = Path(path)
        if not p.exists():
            return f"❌ 文件不存在: {path}"
        wb = openpyxl.load_workbook(p, data_only=True)
        ws = wb[sheet] if sheet else wb.active
        sheet_name = ws.title
        r = f"[工作表: {sheet_name}]\n"
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            r += " | ".join(str(c or "")[:25] for c in row) + "\n"
        return r + f"--- 共{min(max_rows, ws.max_row)}行 ---"
    except ImportError:
        return "❌ 需要 openpyxl: pip install openpyxl"
    except Exception as e:
        return f"❌ Excel错误: {e}"


# ---- 工具5: pdf_extract ----
def pdf_extract(args: Dict[str, Any]) -> str:
    """提取PDF文本"""
    path = args.get("path", "")
    max_chars = int(args.get("max_chars", 3000))
    if not path:
        return "❌ 需提供文件路径"
    try:
        import fitz
        p = Path(path)
        if not p.exists():
            return f"❌ 文件不存在: {path}"
        doc = fitz.open(p)
        text = ""
        for page in doc:
            text += page.get_text()
            if len(text) >= max_chars:
                text = text[:max_chars] + "\n...截断"
                break
        doc.close()
        return text.strip() or "PDF无文本（可能是扫描件）"
    except ImportError:
        return "❌ 需要 pymupdf: pip install pymupdf"
    except Exception as e:
        return f"❌ PDF错误: {e}"


# ---- 工具6: self_heal ----
def self_heal(args: Dict[str, Any] = None) -> str:
    """自检自修：检查关键模块完整性"""
    results = []
    home = Path(os.environ.get("PSPAI_HOME", str(Path(__file__).parent.parent)))
    for d in ["data", "logs"]:
        p = home / d
        if not p.exists():
            p.mkdir(parents=True, exist_ok=True)
            results.append(f"✅ 重建: {d}")
        else:
            results.append(f"✓ {d}")
    for fname in ["config.yaml", "config.json"]:
        cf = home / fname
        if cf.exists():
            results.append(f"✓ {fname} 存在")
            break
    else:
        results.append("⚠️ 无配置文件（将使用环境变量）")
    soul = home / "SOUL.md"
    results.append(f"✓ SOUL.md {'存在' if soul.exists() else '缺失（使用默认提示词）'}")
    try:
        import platform
        if platform.system() == "Linux":
            st = os.statvfs(str(home))
            free = (st.f_bavail * st.f_frsize) / (1024 ** 3)
            results.append(f"✓ 磁盘: {free:.1f}GB")
        else:
            results.append("✓ 磁盘空间（跳过statvfs）")
    except Exception:
        pass
    return "\n".join(results)


# ---- 工具7: tool_doctor ----
def tool_doctor(args: Dict[str, Any] = None) -> str:
    """检查工具健康状态"""
    tools = {
        "git_operations": git_operations, "sql_query": sql_query,
        "csv_read": csv_read, "excel_read": excel_read, "pdf_extract": pdf_extract,
        "self_heal": self_heal, "self_check_gaps": self_check_gaps,
        "tool_doctor": tool_doctor, "decision_why": decision_why,
        "pattern_match": pattern_match, "search_log": search_log,
        "workflow_decompose": workflow_decompose, "ssh_exec": ssh_exec,
        "brother_watch": brother_watch, "ai_search": ai_search,
        "knowledge_query": knowledge_query, "shell_exec": None,
    }
    # 从TOOLS_DEF补充
    import inspect
    tool_name = args.get("tool_name", "") if args else ""
    if tool_name:
        if tool_name not in tools:
            return f"❌ 未知工具: {tool_name}"
        fn = tools[tool_name]
        try:
            sig = inspect.signature(fn)
            return f"✅ {tool_name} 正常（参数: {list(sig.parameters.keys())})"
        except Exception as e:
            return f"❌ {tool_name}: {e}"
    rs = []
    for n, fn in tools.items():
        if fn is None:
            rs.append(f"✅ {n}（内置）")
            continue
        try:
            sig = inspect.signature(fn)
            rs.append(f"✅ {n}（参数: {list(sig.parameters.keys())})")
        except Exception as e:
            rs.append(f"❌ {n}: {e}")
    return "\n".join(rs)


# ---- 工具8: decision_why ----
def decision_why(args: Dict[str, Any]) -> str:
    """决策分析：收益/风险/替代方案"""
    question = args.get("question", "")
    if not question:
        return "❌ 需提供问题"
    return (f"## 决策分析\n\n**问题**: {question}\n\n"
            f"1. 收益分析\n2. 风险评估\n3. 替代方案\n4. 结论建议")


# ---- 工具9: pattern_match ----
def pattern_match(args: Dict[str, Any]) -> str:
    """跨文件模式匹配（类grep）"""
    pattern = args.get("pattern", "")
    path = args.get("path", ".")
    file_glob = args.get("file_glob", "*.py")
    max_results = int(args.get("max_results", 20))
    if not pattern:
        return "❌ 需提供关键词"
    try:
        p = Path(path)
        if not p.exists():
            return f"❌ 路径不存在: {path}"
        rs = []
        for f in p.rglob(file_glob):
            if f.is_file() and f.stat().st_size < 1024 * 1024:
                try:
                    c = f.read_text(encoding="utf-8", errors="replace")
                    for i, line in enumerate(c.split("\n"), 1):
                        if isinstance(pattern, str) and pattern in line and len(rs) < max_results:
                            rs.append(f"{f.relative_to(p)}:{i}: {line.strip()[:100]}")
                except OSError:
                    pass
        if not rs:
            return f"无匹配 '{pattern}'"
        return "\n".join(rs) + (f"\n...共{len(rs)}条" if len(rs) >= max_results else "")
    except Exception as e:
        return f"❌ 失败: {e}"


# ---- 工具10: search_log ----
def search_log(args: Dict[str, Any]) -> str:
    """搜索日志关键词"""
    keyword = args.get("keyword", "")
    log_dir = args.get("log_dir", "~/.pspai/logs")
    lines = int(args.get("lines", 10))
    if not keyword:
        return "❌ 需提供关键词"
    try:
        p = Path(log_dir).expanduser()
        if not p.exists():
            return f"❌ 日志目录不存在: {log_dir}"
        rs = []
        for lf in sorted(p.glob("*.log"), reverse=True):
            try:
                with open(lf, "r", encoding="utf-8", errors="replace") as f:
                    for line in f:
                        if keyword in line:
                            rs.append(f"[{lf.name}] {line.strip()[:200]}")
                            if len(rs) >= lines:
                                break
            except OSError:
                pass
            if len(rs) >= lines:
                break
        return "\n".join(rs) if rs else f"日志中未找到: {keyword}"
    except Exception as e:
        return f"❌ 失败: {e}"


# ---- 工具11: workflow_decompose ----
def workflow_decompose(args: Dict[str, Any]) -> str:
    """复杂目标分解为可执行步骤"""
    goal = args.get("goal", "")
    if not goal:
        return "❌ 需提供目标"
    return (f"## 工作流分解\n\n**目标**: {goal}\n\n"
            f"1. 分析需求\n2. 拆分子任务\n3. 准备资源\n"
            f"4. 分步执行\n5. 验证结果\n6. 总结归档")


# ---- 工具12: ssh_exec ----
HOST_MAP = {
    "M4-1": {"host": os.environ.get("M4_1_HOST", "localhost"), "user": os.environ.get("M4_1_USER", "admin")},
    "M4-2": {"host": os.environ.get("M4_2_HOST", "localhost"), "user": os.environ.get("M4_2_USER", "admin")},
    "DGX": {"host": os.environ.get("DGX_HOST", "127.0.0.1"), "user": os.environ.get("DGX_USER", "admin")},
}

def ssh_exec(args: Dict[str, Any]) -> str:
    """SSH远程执行命令"""
    host = args.get("host", "")
    cmd = args.get("cmd", "")
    if not host or not cmd:
        return "❌ 需提供host和cmd"
    hi = HOST_MAP.get(host.upper(), {})
    if not hi:
        return f"❌ 未知主机: {host}（可用: {', '.join(HOST_MAP.keys())}）"
    if host.upper() == "DGX":
        try:
            import shlex
            r = subprocess.run(shlex.split(cmd), capture_output=True, text=True, timeout=30)
            return (r.stdout or r.stderr)[:2000]
        except subprocess.TimeoutExpired:
            return "❌ 命令超时"
        except Exception as e:
            return f"❌ {e}"
    try:
        import paramiko
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        kp = os.path.expanduser("~/.ssh/id_ed25519")
        pk = paramiko.Ed25519Key.from_private_key_file(kp) if os.path.exists(kp) else None
        ssh.connect(hi["host"], username=hi["user"], pkey=pk, timeout=10)
        _, so, se = ssh.exec_command(cmd, timeout=30)
        out = so.read().decode("utf-8", errors="replace")[:2000]
        err = se.read().decode("utf-8", errors="replace")[:500]
        ssh.close()
        return out or err or "无输出"
    except ImportError:
        return "❌ 需要 paramiko: pip install paramiko"
    except Exception as e:
        return f"❌ SSH错误: {e}"


# ---- 工具13: brother_watch ----
def brother_watch(args: Dict[str, Any] = None) -> str:
    """监控兄弟机器在线状态（ping）"""
    r = "## 兄弟机器状态\n\n"
    for name, info in HOST_MAP.items():
        try:
            res = subprocess.run(["ping", "-c1", "-W2", info["host"]],
                                 capture_output=True, text=True, timeout=5)
            r += f"{'✅' if res.returncode == 0 else '❌'} {name} ({info['host']}) "
            r += "在线\n" if res.returncode == 0 else "离线\n"
        except (subprocess.TimeoutExpired, OSError):
            r += f"⚠️ {name} 超时\n"
    return r


# ══════════════════════════════════════════════
# 第2阶基因：自主规划（NPC2 — Neural Planning Cortex）
# 让小龙人从"一句一答"升级为"接一个任务自动拆成多步做完"
# 碱基类型：T（Skill）
# ══════════════════════════════════════════════

# ---- 工具N1: task_decomposer ----
def task_decomposer(args: Dict[str, Any]) -> str:
    """复杂目标自动拆解为有序子任务"""
    goal = args.get("goal", "")
    if not goal:
        return "❌ 需提供目标"
    context = args.get("context", "")
    
    # 生成结构化的任务分解
    lines = [f"🎯 目标: {goal}"]
    if context:
        lines.append(f"📋 上下文: {context}")
    lines.extend([
        "",
        "┌─ 任务分解 ──────────────────────────────",
        "│",
        "│  Phase 1: 准备与分析",
        "│    Step 1.1  分析需求",
        "│    Step 1.2  收集所需信息/工具",
        "│    Step 1.3  确认前置条件",
        "│",
        "│  Phase 2: 执行阶段",
        "│    Step 2.1  执行核心操作",
        "│    Step 2.2  中间验证",
        "│    Step 2.3  修正/调整",
        "│",
        "│  Phase 3: 交付与验证",
        "│    Step 3.1  结果检查",
        "│    Step 3.2  输出最终结果",
        "└──────────────────────────────────────────",
        "",
        "⚡ 任务状态: 待执行",
        "📊 进度: 0/6 步骤",
    ])
    return "\n".join(lines)


# ---- 工具N2: step_executor ----
def step_executor(args: Dict[str, Any]) -> str:
    """按序列执行多步任务中的一步"""
    step_id = args.get("step_id", "")
    step_desc = args.get("step_desc", "")
    action = args.get("action", "")
    deps = args.get("deps", "")  # 依赖的上一步结果
    
    if not step_id or not action:
        return "❌ 需提供step_id和action"
    
    result = []
    result.append(f"📌 Step {step_id}: {step_desc or action}")
    if deps:
        result.append(f"   🔗 依赖: {deps[:100]}")
    result.append(f"   ⚡ 执行: {action[:200]}")
    result.append(f"   ⏳ 状态: 执行中...")
    
    # 根据action类型处理
    if action.startswith("shell:"):
        cmd = action[6:].strip()
        result.append(f"   🔧 将通过Shell执行: {cmd[:100]}")
        from urllib.parse import quote
        return "\n".join(result) + "\n" + f"GENE://shell_exec?cmd={quote(cmd)}"
    elif action.startswith("tool:"):
        tool_call = action[5:].strip()
        result.append(f"   🔧 将调用工具: {tool_call[:100]}")
        return "\n".join(result) + "\n" + tool_call
    else:
        result.append(f"   ✅ 描述性步骤（无需工具调用）")
        result.append(f"   ✅ Step {step_id} 执行完成")
        return "\n".join(result)


# ---- 工具N3: self_corrector ----
def self_corrector(args: Dict[str, Any]) -> str:
    """某步失败时自动分析根因+重试/换方案"""
    step_id = args.get("step_id", "")
    error_info = args.get("error", "")
    step_desc = args.get("step_desc", "")
    attempt = int(args.get("attempt", "1"))
    
    result = []
    result.append(f"🔍 纠错分析 — Step {step_id}: {step_desc}")
    result.append(f"   ❌ 错误: {error_info[:200]}")
    result.append(f"   🔄 第{attempt}次尝试")
    
    # 错误分类分析
    error_lower = error_info.lower()
    strategies = []
    
    if "timeout" in error_lower or "timed out" in error_lower:
        strategies = ["减少数据量/分步执行", "增加超时时间", "换更快的替代方案"]
    elif "not found" in error_lower or "不存在" in error_lower:
        strategies = ["检查路径是否正确", "使用搜索工具查找", "创建缺失的资源"]
    elif "permission" in error_lower or "denied" in error_lower:
        strategies = ["检查权限设置", "用sudo或管理员身份", "换可访问的路径"]
    elif "network" in error_lower or "connection" in error_lower:
        strategies = ["检查网络连通性", "切换DNS/代理", "等网络恢复后重试"]
    elif "api_key" in error_lower or "unauthorized" in error_lower:
        strategies = ["检查API Key是否正确", "确认API服务状态", "检查余额"]
    elif "parse" in error_lower or "invalid" in error_lower:
        strategies = ["检查输入格式", "修正参数类型", "使用标准格式"]
    else:
        strategies = ["重试一次（可能临时故障）", "换不同的方法实现", "拆分成更小的步骤"]
    
    result.append(f"   💡 策略建议:")
    for i, s in enumerate(strategies[:3], 1):
        result.append(f"     {i}. {s}")
    
    if attempt >= 3:
        result.append(f"\n   ⚠️ 已连续失败{attempt}次，建议换完全不同的方案")
    
    result.append(f"\n   📊 纠错记录: 已尝试{attempt}次")
    
    return "\n".join(result)


# ---- 工具N4: progress_tracker ----
def progress_tracker(args: Dict[str, Any]) -> str:
    """多步执行进度追踪与结果汇总"""
    total = int(args.get("total", 0))
    completed = int(args.get("completed", 0))
    results_summary = args.get("results", "")
    failed = args.get("failed", "")
    
    pct = (completed / total * 100) if total > 0 else 0
    bar_len = 20
    filled = int(bar_len * completed / max(total, 1))
    bar = "█" * filled + "░" * (bar_len - filled)
    
    lines = [
        f"📊 任务执行进度",
        f"",
        f"   {bar}",
        f"   {completed}/{total} 步骤完成 ({pct:.0f}%)",
        f"",
    ]
    
    if results_summary:
        lines.append(f"✅ 已完成步骤结果:")
        for line in results_summary.split("\n")[:5]:
            if line.strip():
                lines.append(f"   • {line.strip()[:100]}")
        lines.append("")
    
    if failed:
        lines.append(f"❌ 失败步骤: {failed[:200]}")
        lines.append("")
    
    if pct >= 100:
        lines.append(f"🎉 全部完成！")
    else:
        remaining = total - completed
        lines.append(f"⏳ 剩余 {remaining} 步")
    
    return "\n".join(lines)


# ══════════════════════════════════════════════
# 第4阶基因：记忆进化（MEG4 — Memory Evolution Gene）
# 从"存对话记录"升级为"真记住你是谁、你的事"
# 碱基类型：T（Skill）
# ══════════════════════════════════════════════

# ---- 工具N5: user_profile ----
def user_profile(args: Dict[str, Any]) -> str:
    """从对话中提取用户偏好/习惯/重要事实并归档为记忆"""
    user_input = args.get("user_input", "")
    ai_response = args.get("ai_response", "")
    
    profile_path = Path(__file__).parent / "data" / "user_profile.json"
    profile = {"preferences": [], "facts": [], "habits": [], "last_updated": time.time()}
    if profile_path.exists():
        try:
            profile = json.loads(profile_path.read_text())
        except:
            pass
    
    now = time.strftime("%Y-%m-%d %H:%M")
    
    # 从对话内容中提取可能的用户信息
    insights = []
    
    # 分析用户输入中的偏好信号
    pref_signals = ["我喜欢", "我要", "我需要", "帮我", "用这个", "习惯"]
    for signal in pref_signals:
        if signal in user_input:
            idx = user_input.find(signal)
            snippet = user_input[idx:idx+60]
            # 去重
            dup = False
            for p in profile["preferences"]:
                if p.get("text","")[:30] == snippet[:30]:
                    dup = True
                    break
            if not dup:
                profile["preferences"].append({"text": snippet, "time": now})
                insights.append(f"📝 记录偏好: {snippet}")
    
    # 事实提取（明显的陈述句）
    fact_signals = ["我是", "我在", "我有", "我用", "我的"]
    for signal in fact_signals:
        if signal in user_input:
            idx = user_input.find(signal)
            snippet = user_input[idx:idx+60]
            profile["facts"].append({"text": snippet, "time": now})
            insights.append(f"📌 记录事实: {snippet}")
    
    # 限制条目数
    for key in ["preferences", "facts", "habits"]:
        if len(profile[key]) > 100:
            profile[key] = profile[key][-100:]
    
    profile["last_updated"] = time.time()
    profile_path.parent.mkdir(exist_ok=True)
    profile_path.write_text(json.dumps(profile, ensure_ascii=False, indent=2))
    
    lines = ["🧠 用户画像更新"]
    if insights:
        for ins in insights[-5:]:
            lines.append(f"  {ins}")
    lines.append(f"  📊 累计: {len(profile['preferences'])}条偏好 · {len(profile['facts'])}条事实")
    if profile["preferences"]:
        lines.append(f"\n📋 最近偏好:")
        for p in profile["preferences"][-3:]:
            lines.append(f"  • {p['text']}")
    if profile["facts"]:
        lines.append(f"\n📋 已知事实:")
        for f in profile["facts"][-3:]:
            lines.append(f"  • {f['text']}")
    
    return "\n".join(lines)


# ---- 工具N6: memory_retrieval ----
def memory_retrieval(args: Dict[str, Any]) -> str:
    """在对话历史和用户画像中搜索相关信息"""
    query = args.get("query", "")
    if not query:
        return "❌ 需提供搜索关键词"
    
    results = []
    now = time.time()
    
    # 1. 搜索对话历史
    from pspai_server import conversation_history
    for char_idx, msgs in conversation_history.items():
        count = 0
        for msg in msgs[-100:]:  # 只看最近100条
            content = msg.get("content", "")
            if query in content:
                role = "👤 用户" if msg.get("role") == "user" else "🤖 AI"
                snippet = content[:200]
                results.append(f"[对话{char_idx}] {role}: {snippet}")
                count += 1
                if count >= 3:
                    break
    
    # 2. 搜索用户画像
    profile_path = Path(__file__).parent / "data" / "user_profile.json"
    if profile_path.exists():
        try:
            profile = json.loads(profile_path.read_text())
            for pref in profile.get("preferences", []):
                if query in pref.get("text", ""):
                    results.append(f"[偏好] {pref['text']} ({pref.get('time','')})")
            for fact in profile.get("facts", []):
                if query in fact.get("text", ""):
                    results.append(f"[事实] {fact['text']} ({fact.get('time','')})")
        except:
            pass
    
    if not results:
        return f"未找到与 '{query}' 相关记忆"
    
    lines = [f"🔍 记忆检索: '{query}'", f"  找到 {len(results)} 条相关记忆", ""]
    for i, r in enumerate(results[:10], 1):
        lines.append(f"  {i}. {r[:150]}")
    
    return "\n".join(lines)


# ---- 工具N7: knowledge_accumulate ----
def knowledge_accumulate(args: Dict[str, Any]) -> str:
    """从对话中提取知识片段并聚合为结构化学识"""
    topic = args.get("topic", "")
    new_knowledge = args.get("content", "")
    source = args.get("source", "对话")
    
    knowledge_path = Path(__file__).parent / "data" / "knowledge.json"
    knowledge = []
    if knowledge_path.exists():
        try:
            knowledge = json.loads(knowledge_path.read_text())
        except:
            pass
    
    now = time.strftime("%Y-%m-%d %H:%M")
    
    if new_knowledge:
        # 添加新知识条目
        entry = {
            "text": new_knowledge[:500],
            "topic": topic or "通用",
            "source": source,
            "time": now,
            "access_count": 0,
        }
        # 去重：如果内容高度相似就不加
        if not any(e["text"][:50] == new_knowledge[:50] for e in knowledge[-20:]):
            knowledge.append(entry)
    
    # 按主题聚合统计
    topics = {}
    for k in knowledge:
        t = k.get("topic", "通用")
        if t not in topics:
            topics[t] = {"count": 0, "recent": []}
        topics[t]["count"] += 1
        topics[t]["recent"].append(k)
    
    lines = [f"📚 学识积累"]
    if new_knowledge:
        lines.append(f"  ✅ 已添加: {new_knowledge[:100]}")
    lines.append(f"  📊 总学识: {len(knowledge)} 条")
    lines.append(f"")
    lines.append(f"📋 主题分布:")
    for t, info in sorted(topics.items(), key=lambda x: -x[1]["count"])[:5]:
        lines.append(f"  • {t}: {info['count']}条")
        if info["recent"]:
            last = info["recent"][-1]
            lines.append(f"    最近: {last['text'][:80]} ({last.get('time','')})")
    
    # 持久化
    if len(knowledge) > 500:
        knowledge = knowledge[-500:]
    
    knowledge_path.parent.mkdir(exist_ok=True)
    knowledge_path.write_text(json.dumps(knowledge, ensure_ascii=False, indent=2))
    
    return "\n".join(lines)


# ---- 工具N8: cross_session ----
def cross_session(args: Dict[str, Any]) -> str:
    """跨会话记忆——汇总所有已知信息以保持连续性"""
    action = args.get("action", "summary")
    
    profile_path = Path(__file__).parent / "data" / "user_profile.json"
    knowledge_path = Path(__file__).parent / "data" / "knowledge.json"
    
    summary_lines = ["🌐 跨会话记忆摘要"]
    
    if action == "summary" or action == "all":
        # 用户画像摘要
        profile = {}
        if profile_path.exists():
            try:
                profile = json.loads(profile_path.read_text())
            except:
                pass
        
        facts = profile.get("facts", [])
        prefs = profile.get("preferences", [])
        if facts:
            summary_lines.append(f"\n👤 关于用户:")
            for f in facts[-5:]:
                summary_lines.append(f"  • {f.get('text','')[:100]}")
        if prefs:
            summary_lines.append(f"\n❤️ 用户偏好:")
            for p in prefs[-5:]:
                summary_lines.append(f"  • {p.get('text','')[:100]}")
        
        # 学识摘要
        knowledge = []
        if knowledge_path.exists():
            try:
                knowledge = json.loads(knowledge_path.read_text())
            except:
                pass
        
        if knowledge:
            summary_lines.append(f"\n📚 已积累学识 ({len(knowledge)}条):")
            for k in knowledge[-5:]:
                summary_lines.append(f"  • [{k.get('topic','通用')}] {k.get('text','')[:80]}")
        
        # 会话数量
        try:
            from pspai_server import conversation_history
            total_sessions = len(conversation_history)
            total_msgs = sum(len(msgs) for msgs in conversation_history.values())
            summary_lines.append(f"\n📊 统计: {total_sessions}个对话 · {total_msgs}条消息")
        except:
            pass
    
    elif action == "clear":
        # 重置跨会话记忆（保留最新10条）
        if profile_path.exists():
            try:
                profile = json.loads(profile_path.read_text())
                for k in profile:
                    if isinstance(profile[k], list):
                        profile[k] = profile[k][-10:]
                profile_path.write_text(json.dumps(profile, ensure_ascii=False, indent=2))
                summary_lines.append(f"  🗑️ 已清理画像（保留最近10条）")
            except:
                pass
        summary_lines.append(f"\n✅ 跨会话记忆已重置")
    
    return "\n".join(summary_lines)


# ---- 工具14: ai_search ----
def ai_search(args: Dict[str, Any]) -> str:
    """搜索互联网（DDG优先，超时自动回退）"""
    query = args.get("query", "")
    max_results = int(args.get("max_results", 5))
    if not query:
        return "❌ 需提供搜索词"
    try:
        import requests
        headers = {"User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36"}
        # 尝试多个搜索引擎
        urls = [
            f"https://html.duckduckgo.com/html/?q={query.replace(' ', '+')}",
            f"https://www.baidu.com/s?wd={query.replace(' ', '+')}",
        ]
        resp = None
        last_err = ""
        for url in urls:
            try:
                resp = requests.get(url, headers=headers, timeout=8)
                if resp.status_code == 200:
                    break
            except Exception as e:
                last_err = str(e)[:50]
                continue
        if resp is None:
            return f"❌ 搜索失败（网络不可达: {last_err}）"
        # 提取标题链接
        titles = re.findall(r'<a[^>]*>(.*?)</a>', resp.text)
        results = [t.strip() for t in titles if t.strip() and len(t.strip()) > 5 and '<' not in t]
        if not results:
            results = re.findall(r'<h[1-3][^>]*>(.*?)</h[1-3]>', resp.text)
        results = [re.sub(r'<[^>]+>', '', r).strip() for r in results[:max_results] if r.strip()]
        if not results:
            return f"搜索 '{query}' 无结果"
        return "搜索结果:\n\n" + "\n\n".join(f"• {r[:200]}" for r in results)
    except ImportError:
        return "❌ 需要 requests: pip install requests"


# ---- 工具15: knowledge_query ----
def knowledge_query(args: Dict[str, Any]) -> str:
    """查询本地知识库"""
    question = args.get("question", "")
    knowledge_dir = args.get("knowledge_dir", str(Path.home() / "桌面" / "平行时空人类、AI知识大全库"))
    if not question:
        return "❌ 需提供问题"
    try:
        kd = Path(knowledge_dir)
        if not kd.exists():
            return f"❌ 知识库不存在: {knowledge_dir}"
        results = []
        for f in kd.rglob("*.md"):
            if f.stat().st_size > 1024 * 500:
                continue
            try:
                c = f.read_text(encoding="utf-8", errors="replace")
                if question[:10] in c:
                    rel = f.relative_to(kd)
                    results.append(f"[{rel}]\n{c[:500].strip()}")
                    if len(results) >= 3:
                        break
            except Exception:
                pass
        if not results:
            return "知识库中未找到相关信息"
        return "\n\n---\n\n".join(results)
    except Exception as e:
        return f"❌ 知识查询失败: {e}"


# ---- 工具16: self_check_gaps ----
def self_check_gaps(args: Dict[str, Any] = None) -> str:
    """检查工具能力覆盖缺口"""
    # 在函数内部延迟导入，避免循环导入
    from pspai_server import DNAREGISTRY
    names = list(DNAREGISTRY.genes.keys())
    desc_all = " ".join(g.description for g in DNAREGISTRY.genes.values()).lower()
    standard = {
        "文件": ["读", "写", "搜索", "编辑"], "网络": ["搜索", "抓取", "API"],
        "执行": ["Shell", "代码"], "代码": ["Git", "SQL"],
        "数据": ["CSV", "Excel", "PDF"], "系统": ["SSH", "自检", "监控"],
        "搜索": ["AI搜索", "知识库"],
    }
    own = {
        "文件": ["读", "写", "搜索", "编辑"], "网络": [],
        "执行": ["Shell"], "代码": ["Git", "SQL"],
        "数据": ["CSV", "Excel", "PDF"], "系统": ["SSH", "自检", "监控"],
        "搜索": ["AI搜索", "知识库"],
    }
    result = ""
    for cat, items in standard.items():
        has_items = own.get(cat, [])
        found_from_desc = [s for s in items if s.lower() in desc_all and s not in has_items]
        all_have = has_items + found_from_desc
        missing = [s for s in items if s not in all_have]
        if missing:
            result += f"⚠️ {cat}: 缺{'/'.join(missing)}（有{'/'.join(all_have) if all_have else '无'}）\n"
        else:
            result += f"✅ {cat}: 完整\n"
    return result or "✅ 全部完整"


# ---- 工具17: shell_exec（在pspai_server.py中定义）----

# ══════════════════════════════════════════════
# 注册器
# ══════════════════════════════════════════════

SHELL_EXEC_FN = None  # pspai_server.py会在main()中设置

TOOLS_DEF = {
    "git_operations": {"fn": git_operations,
        "desc": "Git操作: status/add/commit/push/pull/log/diff",
        "params": {"type":"object","properties":{
            "action":{"type":"string","description":"操作类型"},
            "repo_path":{"type":"string","description":"仓库路径","default":"."},
            "message":{"type":"string","description":"commit消息","default":""},
            "branch":{"type":"string","description":"分支名","default":""},
            "remote":{"type":"string","description":"远程仓库","default":"origin"}},
            "required":["action"]}},
    "sql_query": {"fn": sql_query,
        "desc": "执行SQL查询（SQLite）",
        "params": {"type":"object","properties":{
            "query":{"type":"string","description":"SQL语句"},
            "db_path":{"type":"string","description":"数据库路径","default":""}},
            "required":["query"]}},
    "csv_read": {"fn": csv_read,
        "desc": "读取CSV文件",
        "params": {"type":"object","properties":{
            "path":{"type":"string","description":"文件路径"},
            "max_rows":{"type":"integer","description":"最大行数","default":20}},
            "required":["path"]}},
    "excel_read": {"fn": excel_read,
        "desc": "读取Excel文件",
        "params": {"type":"object","properties":{
            "path":{"type":"string","description":"文件路径"},
            "sheet":{"type":"string","description":"工作表名","default":""},
            "max_rows":{"type":"integer","description":"最大行数","default":20}},
            "required":["path"]}},
    "pdf_extract": {"fn": pdf_extract,
        "desc": "提取PDF文本",
        "params": {"type":"object","properties":{
            "path":{"type":"string","description":"文件路径"},
            "max_chars":{"type":"integer","description":"最大字符数","default":3000}},
            "required":["path"]}},
    "self_heal": {"fn": self_heal,
        "desc": "自检自修：检查关键模块完整性",
        "params": {"type":"object","properties":{}}},
    "tool_doctor": {"fn": tool_doctor,
        "desc": "检查工具健康状态",
        "params": {"type":"object","properties":{
            "tool_name":{"type":"string","description":"工具名（空则全检）","default":""}}}},
    "decision_why": {"fn": decision_why,
        "desc": "决策分析：收益/风险/替代方案",
        "params": {"type":"object","properties":{
            "question":{"type":"string","description":"需要分析的问题"}},
            "required":["question"]}},
    "pattern_match": {"fn": pattern_match,
        "desc": "跨文件模式匹配（类grep）",
        "params": {"type":"object","properties":{
            "pattern":{"type":"string","description":"关键词"},
            "path":{"type":"string","description":"搜索路径","default":"."},
            "file_glob":{"type":"string","description":"文件通配符","default":"*.py"},
            "max_results":{"type":"integer","description":"最大结果数","default":20}},
            "required":["pattern"]}},
    "search_log": {"fn": search_log,
        "desc": "搜索日志关键词",
        "params": {"type":"object","properties":{
            "keyword":{"type":"string","description":"关键词"},
            "log_dir":{"type":"string","description":"日志目录","default":"~/.pspai/logs"},
            "lines":{"type":"integer","description":"返回行数","default":10}},
            "required":["keyword"]}},
    "workflow_decompose": {"fn": workflow_decompose,
        "desc": "复杂目标分解为可执行步骤",
        "params": {"type":"object","properties":{
            "goal":{"type":"string","description":"目标描述"}},
            "required":["goal"]}},
    "ssh_exec": {"fn": ssh_exec,
        "desc": "SSH远程执行命令: host=M4-1/M4-2/DGX",
        "params": {"type":"object","properties":{
            "host":{"type":"string","description":"目标主机: M4-1/M4-2/DGX"},
            "cmd":{"type":"string","description":"要执行的命令"}},
            "required":["host","cmd"]}},
    "brother_watch": {"fn": brother_watch,
        "desc": "监控兄弟机器在线状态（ping）",
        "params": {"type":"object","properties":{}}},
    "ai_search": {"fn": ai_search,
        "desc": "搜索互联网",
        "params": {"type":"object","properties":{
            "query":{"type":"string","description":"搜索词"},
            "max_results":{"type":"integer","description":"结果数","default":5}},
            "required":["query"]}},
    "knowledge_query": {"fn": knowledge_query,
        "desc": "查询本地知识库",
        "params": {"type":"object","properties":{
            "question":{"type":"string","description":"问题"},
            "knowledge_dir":{"type":"string","description":"知识库路径","default":""}},
            "required":["question"]}},
    "self_check_gaps": {"fn": self_check_gaps,
        "desc": "检查工具能力覆盖缺口",
        "params": {"type":"object","properties":{}}},
    # ── 第2阶基因：自主规划（NPC2） ──
    "task_decomposer": {"fn": task_decomposer,
        "desc": "复杂目标自动拆解为有序子任务",
        "params": {"type":"object","properties":{
            "goal":{"type":"string","description":"要完成的目标"},
            "context":{"type":"string","description":"背景信息","default":""}},
            "required":["goal"]}},
    "step_executor": {"fn": step_executor,
        "desc": "按序列执行多步任务中的一步",
        "params": {"type":"object","properties":{
            "step_id":{"type":"string","description":"步骤编号"},
            "step_desc":{"type":"string","description":"步骤描述","default":""},
            "action":{"type":"string","description":"执行动作: shell:命令 / tool:工具调用"},
            "deps":{"type":"string","description":"上一步结果依赖","default":""}},
            "required":["step_id","action"]}},
    "self_corrector": {"fn": self_corrector,
        "desc": "步骤失败时自动分析根因+重试/换方案",
        "params": {"type":"object","properties":{
            "step_id":{"type":"string","description":"失败的步骤编号"},
            "error":{"type":"string","description":"错误信息"},
            "step_desc":{"type":"string","description":"步骤描述","default":""},
            "attempt":{"type":"string","description":"已尝试次数","default":"1"}},
            "required":["step_id","error"]}},
    "progress_tracker": {"fn": progress_tracker,
        "desc": "多步执行进度追踪与结果汇总",
        "params": {"type":"object","properties":{
            "total":{"type":"string","description":"总步骤数","default":"0"},
            "completed":{"type":"string","description":"已完成步骤数","default":"0"},
            "results":{"type":"string","description":"步骤结果汇总","default":""},
            "failed":{"type":"string","description":"失败的步骤","default":""}}}},

    # ── 第4阶基因：记忆进化（MEG4） ──
    "user_profile": {"fn": user_profile,
        "desc": "从对话提取用户偏好/习惯/重要事实并归档",
        "params": {"type":"object","properties":{
            "user_input":{"type":"string","description":"用户输入文本"},
            "ai_response":{"type":"string","description":"AI回复文本","default":""}},
            "required":["user_input"]}},
    "memory_retrieval": {"fn": memory_retrieval,
        "desc": "在对话历史和画像中搜索相关记忆",
        "params": {"type":"object","properties":{
            "query":{"type":"string","description":"搜索关键词"}},
            "required":["query"]}},
    "knowledge_accumulate": {"fn": knowledge_accumulate,
        "desc": "从对话提取知识并聚合为结构化学识",
        "params": {"type":"object","properties":{
            "topic":{"type":"string","description":"知识主题","default":""},
            "content":{"type":"string","description":"知识内容"},
            "source":{"type":"string","description":"来源","default":"对话"}},
            "required":["content"]}},
    "cross_session": {"fn": cross_session,
        "desc": "跨会话记忆——汇总所有已知信息保持连续性",
        "params": {"type":"object","properties":{
            "action":{"type":"string","description":"操作: summary/clear","default":"summary"}}}},
    "loop_worker_exec": {"fn": None,  # 延迟绑定，见文件末尾
        "desc": "【第6阶·Loop引擎】Worker执行——只干活不判断好坏。写文件/查服务/跑命令",
        "params": {"type":"object","properties":{
            "task_type":{"type":"string","description":"任务类型: write_file/check_service/run_command"},
            "args":{"type":"object","description":"任务参数, 如 {\"path\":\"/tmp/x.txt\",\"content\":\"hello\"}"}},
            "required":["task_type","args"]}},
    "loop_reviewer_verify": {"fn": None,  # 延迟绑定
        "desc": "【第6阶·Loop引擎】Reviewer验收——独立检查Worker产出。文件存在/内容完整/服务运行/命令验证",
        "params": {"type":"object","properties":{
            "task_type":{"type":"string","description":"验收类型: file_exists/service_running/command_result"},
            "args":{"type":"object","description":"验收参数, 如 {\"path\":\"/tmp/x.txt\",\"min_bytes\":10,\"keyword\":\"hello\"}"}},
            "required":["task_type","args"]}},
    "loop_full_cycle": {"fn": None,  # 延迟绑定
        "desc": "【第6阶·Loop引擎】全闭环——Worker执行+Reviewer验收一次性完成，返回通过/失败和详情",
        "params": {"type":"object","properties":{
            "task_type":{"type":"string","description":"任务类型: write_file/check_service"},
            "worker_args":{"type":"object","description":"Worker参数"},
            "reviewer_args":{"type":"object","description":"Reviewer验收参数"}},
            "required":["task_type","worker_args","reviewer_args"]}},
    "loop_history": {"fn": None,  # 延迟绑定
        "desc": "【第6阶·Loop引擎】查看闭环执行历史摘要",
        "params": {"type":"object","properties":{
            "limit":{"type":"integer","description":"最近N条记录","default":5}}}},

}

def register_all(register_fn):
    """注册所有工具到pspai_server的DNAREGISTRY"""
    for name, t in TOOLS_DEF.items():
        register_fn(name, t["desc"], t["params"], t["fn"])
    print(f"✅ PSPAI工具集已注册: {len(TOOLS_DEF)}个")


def get_self_check_gaps():
    """外部调用self_check_gaps"""
    return self_check_gaps({})



# ══════════════════════════════════════════════
# 第6阶：Loop闭环基因 — Worker/Reviewer 分离
# ══════════════════════════════════════════════

_LOOP_HISTORY = []  # 闭环执行记录

def loop_worker_exec(args: dict) -> str:
    """Worker执行——只干活不判断好坏"""
    task_type = args.get("task_type", "")
    w_args = args.get("args", {})
    
    if task_type == "write_file":
        path = w_args.get("path", "")
        content = w_args.get("content", "")
        if not path:
            return json.dumps({"success": False, "error": "path required"})
        os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
        return json.dumps({"success": True, "path": path, "bytes": len(content)})
    
    elif task_type == "check_service":
        name = w_args.get("name", "")
        if not name:
            return json.dumps({"success": False, "error": "name required"})
        try:
            result = subprocess.run(["pgrep", "-f", name], capture_output=True, text=True, timeout=10)
            pids = [int(p) for p in result.stdout.strip().split("\n") if p.strip()]
            return json.dumps({"success": True, "running": len(pids) > 0, "pids": pids[:5]})
        except Exception as e:
            return json.dumps({"success": False, "error": str(e)})
    
    elif task_type == "run_command":
        cmd = w_args.get("cmd", "")
        if not cmd:
            return json.dumps({"success": False, "error": "cmd required"})
        try:
            result = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=30)
            return json.dumps({
                "success": result.returncode == 0,
                "exit_code": result.returncode,
                "stdout": result.stdout[-1000:],
                "stderr": result.stderr[-500:],
            })
        except Exception as e:
            return json.dumps({"success": False, "error": str(e)})
    
    else:
        return json.dumps({"success": False, "error": f"Unknown task_type: {task_type}"})


def loop_reviewer_verify(args: dict) -> str:
    """Reviewer验收——独立检查Worker产出"""
    task_type = args.get("task_type", "")
    r_args = args.get("args", {})
    
    if task_type == "file_exists":
        path = r_args.get("path", "")
        min_bytes = r_args.get("min_bytes", 1)
        keyword = r_args.get("keyword", "")
        
        if not path or not os.path.exists(path):
            return json.dumps({"pass": False, "reason": f"文件不存在: {path}"})
        
        size = os.path.getsize(path)
        if size < min_bytes:
            return json.dumps({"pass": False, "reason": f"文件太小: {size} < {min_bytes} bytes", "size": size})
        
        if keyword:
            with open(path, encoding="utf-8", errors="replace") as f:
                content = f.read()
            if keyword not in content:
                return json.dumps({"pass": False, "reason": f"缺少关键词: {keyword}", "size": size})
        
        return json.dumps({"pass": True, "size": size, "path": path})
    
    elif task_type == "service_running":
        name = r_args.get("name", "")
        expected = r_args.get("expected_running", True)
        
        try:
            result = subprocess.run(["pgrep", "-f", name], capture_output=True, text=True, timeout=10)
            pids = [int(p) for p in result.stdout.strip().split("\n") if p.strip()]
            is_running = len(pids) > 0
            
            if is_running != expected:
                status = "运行中" if is_running else "已停止"
                want = "应运行" if expected else "应停止"
                return json.dumps({"pass": False, "reason": f"服务状态不符: {status}, {want}", "pids": pids[:3]})
            return json.dumps({"pass": True, "running": is_running, "pids": pids[:3]})
        except Exception as e:
            return json.dumps({"pass": False, "reason": f"检查异常: {e}"})
    
    elif task_type == "command_result":
        verify_cmd = r_args.get("verify_cmd", "")
        expected_in = r_args.get("expected_in_output", "")
        
        if not verify_cmd:
            return json.dumps({"pass": False, "reason": "no verify_cmd"})
        try:
            result = subprocess.run(verify_cmd, shell=True, capture_output=True, text=True, timeout=30)
            output = result.stdout + result.stderr
            if expected_in and expected_in not in output:
                return json.dumps({"pass": False, "reason": f"验证输出中未找到: {expected_in}"})
            return json.dumps({"pass": True, "verify_output": output[:500]})
        except Exception as e:
            return json.dumps({"pass": False, "reason": f"验证命令异常: {e}"})
    
    else:
        return json.dumps({"pass": False, "reason": f"Unknown review type: {task_type}"})


def loop_full_cycle(args: dict) -> str:
    """全闭环——Worker执行+Reviewer验收一次性完成"""
    task_type = args.get("task_type", "")
    w_args = args.get("worker_args", {})
    r_args = args.get("reviewer_args", {})
    
    t0 = time.time()
    
    # 类型映射：Worker任务类型 → Reviewer验收类型
    review_type_map = {
        "write_file": "file_exists",
        "check_service": "service_running",
        "run_command": "command_result",
    }
    review_type = review_type_map.get(task_type, task_type)
    
    # Worker
    w_result = loop_worker_exec({"task_type": task_type, "args": w_args})
    w_data = json.loads(w_result)
    
    if not w_data.get("success", False):
        return json.dumps({
            "passed": False,
            "stage": "worker",
            "worker_result": w_data,
            "duration": round(time.time() - t0, 2),
        })
    
    # Reviewer（自动把Worker产出路径传给Reviewer）
    r_input = dict(r_args)
    if "path" not in r_input and w_data.get("path"):
        r_input["path"] = w_data["path"]
    if "name" not in r_input and w_data.get("name"):
        r_input["name"] = w_data["name"]
    
    r_result = loop_reviewer_verify({"task_type": review_type, "args": r_input})
    r_data = json.loads(r_result)
    
    passed = r_data.get("pass", False)
    duration = round(time.time() - t0, 2)
    
    record = {
        "task_type": task_type,
        "passed": passed,
        "duration": duration,
        "timestamp": datetime.now().isoformat(),
    }
    _LOOP_HISTORY.append(record)
    if len(_LOOP_HISTORY) > 100:
        _LOOP_HISTORY.pop(0)
    
    return json.dumps({
        "passed": passed,
        "worker_result": w_data,
        "reviewer_result": r_data,
        "duration": duration,
    })


def loop_history(args: dict) -> str:
    """查看Loop闭环执行历史"""
    limit = args.get("limit", 5)
    records = _LOOP_HISTORY[-limit:]
    total = len(_LOOP_HISTORY)
    passed = sum(1 for r in _LOOP_HISTORY if r["passed"])
    return json.dumps({
        "total": total,
        "passed": passed,
        "failed": total - passed,
        "pass_rate": f"{passed/max(total,1)*100:.0f}%",
        "recent": records,
    })

def _bind_loop_tools():
    """延迟绑定Loop工具函数（TOOLS_DEF在函数定义之前，所以fn必须延迟）"""
    global TOOLS_DEF
    TOOLS_DEF["loop_worker_exec"]["fn"] = loop_worker_exec
    TOOLS_DEF["loop_reviewer_verify"]["fn"] = loop_reviewer_verify
    TOOLS_DEF["loop_full_cycle"]["fn"] = loop_full_cycle
    TOOLS_DEF["loop_history"]["fn"] = loop_history

_bind_loop_tools()
